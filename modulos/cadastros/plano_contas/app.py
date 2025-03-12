# coding: utf-8
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, send_file
import os
import logging
from datetime import datetime
import mysql.connector
from mysql.connector import Error
from utils.auth import verificar_permissao
import pandas as pd
from werkzeug.utils import secure_filename
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Configuração de logging
log_dir = os.path.join(os.path.dirname(os.path.dirname(
    os.path.dirname(os.path.abspath(__file__)))), 'logs')
os.makedirs(log_dir, exist_ok=True)
log_file = os.path.join(log_dir, 'plano_contas.log')

logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger('plano_contas')

# Configuração para upload de arquivos
UPLOAD_FOLDER = os.path.join(os.path.dirname(
    os.path.abspath(__file__)), 'uploads')
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

# Criar diretório de uploads se não existir
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Verificar extensão de arquivo permitida


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# Criar o Blueprint
mod_plano_contas = Blueprint('plano_contas', __name__,
                             template_folder='templates')


def get_db_connection():
    try:
        connection = mysql.connector.connect(
            host=os.environ.get('DB_HOST', 'localhost'),
            database=os.environ.get('DB_NAME', 'sistema_solicitacoes'),
            user=os.environ.get('DB_USER', 'root'),
            password=os.environ.get('DB_PASSWORD', 'sua_senha')
        )
        return connection
    except Error as e:
        logger.error(f"Erro ao conectar ao MySQL: {e}")
        return None

# Rota principal - redirecionamento para lista de planos de conta


@mod_plano_contas.route('/')
def index():
    """Redireciona para a página de listagem de planos de conta"""
    return redirect(url_for('cadastros.plano_contas.listar'))

# Rota para listar planos de conta


@mod_plano_contas.route('/listar')
def listar():
    """Exibe a lista de planos de conta cadastrados"""
    # Verificar autenticação
    if 'usuario_id' not in session:
        flash('Você precisa fazer login para acessar esta página', 'warning')
        return redirect(url_for('login'))

    # Verificar permissão
    pode_editar = verificar_permissao('cadastros.plano_contas.editar')

    # Parâmetros de filtro
    numero = request.args.get('numero', '')
    indice = request.args.get('indice', '')
    descricao = request.args.get('descricao', '')
    status = request.args.get('status', '')

    planos = []
    connection = get_db_connection()

    if connection:
        try:
            cursor = connection.cursor(dictionary=True)

            # Montagem dinâmica da consulta SQL com base nos filtros
            query = "SELECT * FROM plano_contas WHERE 1=1"
            params = []

            if numero:
                query += " AND numero LIKE %s"
                params.append(f"%{numero}%")

            if indice:
                query += " AND indice LIKE %s"
                params.append(f"%{indice}%")

            if descricao:
                query += " AND descricao LIKE %s"
                params.append(f"%{descricao}%")

            if status and status.isdigit():
                query += " AND ativo = %s"
                params.append(int(status))

            query += " ORDER BY sequencia, numero"

            cursor.execute(query, params)
            planos = cursor.fetchall()
            cursor.close()
        except Exception as e:
            logger.error(f"Erro ao buscar planos de conta: {str(e)}")
            flash(f'Erro ao buscar planos de conta: {str(e)}', 'danger')
        finally:
            connection.close()

    return render_template('cadastros/plano_contas/listar.html',
                           planos=planos,
                           pode_editar=pode_editar,
                           request=request)

# Rota para cadastrar novo plano de conta


@mod_plano_contas.route('/novo', methods=['GET', 'POST'])
def novo():
    """Exibe o formulário para cadastro de novo plano de conta e processa o envio"""
    # Verificar autenticação
    if 'usuario_id' not in session:
        flash('Você precisa fazer login para acessar esta página', 'warning')
        return redirect(url_for('login'))

    # Verificar permissão
    if not verificar_permissao('cadastros.plano_contas.editar'):
        flash('Você não tem permissão para cadastrar planos de conta', 'danger')
        return redirect(url_for('cadastros.plano_contas.listar'))

    form = {'errors': {}}

    if request.method == 'POST':
        numero = request.form.get('numero')
        indice = request.form.get('indice')
        descricao = request.form.get('descricao')
        sequencia = request.form.get('sequencia')

        # Validações básicas
        if not numero:
            form['errors']['numero'] = 'O número é obrigatório'
        if not indice:
            form['errors']['indice'] = 'O índice é obrigatório'
        if not descricao:
            form['errors']['descricao'] = 'A descrição é obrigatória'

        # Garantir que sequência seja um valor numérico
        try:
            sequencia = int(sequencia) if sequencia else 0
        except ValueError:
            form['errors']['sequencia'] = 'A sequência deve ser um número inteiro'
            sequencia = 0

        # Se não houver erros, prossegue com o cadastro
        if not form['errors']:
            connection = get_db_connection()
            if connection:
                try:
                    cursor = connection.cursor()
                    cursor.execute("""
                        INSERT INTO plano_contas (numero, indice, descricao, sequencia, ativo, criado_em)
                        VALUES (%s, %s, %s, %s, TRUE, NOW())
                    """, (numero, indice, descricao, sequencia))

                    connection.commit()
                    flash('Plano de conta cadastrado com sucesso!', 'success')
                    return redirect(url_for('cadastros.plano_contas.listar'))

                except Exception as e:
                    connection.rollback()
                    logger.error(f"Erro ao cadastrar plano de conta: {str(e)}")
                    flash(
                        f'Erro ao cadastrar plano de conta: {str(e)}', 'danger')
                finally:
                    cursor.close()
                    connection.close()

    return render_template('cadastros/plano_contas/form.html', form=form)

# Rota para visualizar plano de conta


@mod_plano_contas.route('/visualizar/<int:id>')
def visualizar(id):
    """Exibe os detalhes de um plano de conta específico"""
    # Verificar autenticação
    if 'usuario_id' not in session:
        flash('Você precisa fazer login para acessar esta página', 'warning')
        return redirect(url_for('login'))

    # Verificar permissão de edição para ações
    pode_editar = verificar_permissao('cadastros.plano_contas.editar')

    # Verificar se o ID é válido
    try:
        id_plano = int(id)
    except (ValueError, TypeError):
        flash('ID de plano de conta inválido', 'danger')
        return redirect(url_for('cadastros.plano_contas.listar'))

    plano = None
    connection = get_db_connection()

    if connection:
        try:
            cursor = connection.cursor(dictionary=True)
            cursor.execute(
                "SELECT * FROM plano_contas WHERE id = %s", (id_plano,))
            plano = cursor.fetchone()
            cursor.close()

            if not plano:
                flash('Plano de conta não encontrado', 'warning')
                return redirect(url_for('cadastros.plano_contas.listar'))

        except Exception as e:
            logger.error(f"Erro ao buscar plano de conta: {str(e)}")
            flash(f'Erro ao buscar plano de conta: {str(e)}', 'danger')
        finally:
            connection.close()

    return render_template('cadastros/plano_contas/visualizar.html',
                           plano=plano,
                           pode_editar=pode_editar)

# Rota para editar plano de conta


@mod_plano_contas.route('/editar/<int:id>', methods=['GET', 'POST'])
def editar(id):
    """Exibe o formulário para edição de plano de conta e processa o envio"""
    # Verificar autenticação
    if 'usuario_id' not in session:
        flash('Você precisa fazer login para acessar esta página', 'warning')
        return redirect(url_for('login'))

    # Verificar permissão
    if not verificar_permissao('cadastros.plano_contas.editar'):
        flash('Você não tem permissão para editar planos de conta', 'danger')
        return redirect(url_for('cadastros.plano_contas.listar'))

    # Verificar se o ID é válido
    try:
        id_plano = int(id)
    except (ValueError, TypeError):
        flash('ID de plano de conta inválido', 'danger')
        return redirect(url_for('cadastros.plano_contas.listar'))

    plano = None
    form = {'errors': {}}
    connection = get_db_connection()

    if connection:
        try:
            cursor = connection.cursor(dictionary=True)

            if request.method == 'POST':
                numero = request.form.get('numero')
                indice = request.form.get('indice')
                descricao = request.form.get('descricao')
                sequencia = request.form.get('sequencia')
                status = request.form.get('status', '1')

                # Validações básicas
                if not numero:
                    form['errors']['numero'] = 'O número é obrigatório'
                if not indice:
                    form['errors']['indice'] = 'O índice é obrigatório'
                if not descricao:
                    form['errors']['descricao'] = 'A descrição é obrigatória'

                # Garantir que sequência e status sejam valores numéricos
                try:
                    sequencia = int(sequencia) if sequencia else 0
                except ValueError:
                    form['errors']['sequencia'] = 'A sequência deve ser um número inteiro'
                    sequencia = 0

                # Garantir que status seja um valor numérico válido
                status_valor = 1  # valor padrão
                if status and status.isdigit():
                    status_valor = int(status)

                # Se não houver erros, prossegue com a atualização
                if not form['errors']:
                    cursor.execute("""
                        UPDATE plano_contas 
                        SET numero = %s, indice = %s, descricao = %s, sequencia = %s, ativo = %s, atualizado_em = NOW()
                        WHERE id = %s
                    """, (numero, indice, descricao, sequencia, status_valor, id_plano))

                    connection.commit()
                    flash('Plano de conta atualizado com sucesso!', 'success')
                    return redirect(url_for('cadastros.plano_contas.visualizar', id=id_plano))
            else:
                # Buscar dados do plano de conta para exibir no formulário
                cursor.execute(
                    "SELECT * FROM plano_contas WHERE id = %s", (id_plano,))
                plano = cursor.fetchone()

                if not plano:
                    flash('Plano de conta não encontrado', 'warning')
                    return redirect(url_for('cadastros.plano_contas.listar'))

        except Exception as e:
            if request.method == 'POST':
                connection.rollback()
            logger.error(f"Erro ao processar plano de conta: {str(e)}")
            flash(f'Erro ao processar plano de conta: {str(e)}', 'danger')
        finally:
            cursor.close()
            connection.close()

    return render_template('cadastros/plano_contas/form.html', plano=plano, form=form)

# Rota para alternar status do plano de conta (ativar/desativar)


@mod_plano_contas.route('/alternar-status/<int:id>')
def alternar_status(id):
    """Alterna o status (ativo/inativo) de um plano de conta"""
    # Verificar autenticação
    if 'usuario_id' not in session:
        flash('Você precisa fazer login para acessar esta página', 'warning')
        return redirect(url_for('login'))

    # Verificar permissão
    if not verificar_permissao('cadastros.plano_contas.editar'):
        flash('Você não tem permissão para alterar o status de planos de conta', 'danger')
        return redirect(url_for('cadastros.plano_contas.listar'))

    # Verificar se o ID é válido
    try:
        id_plano = int(id)
    except (ValueError, TypeError):
        flash('ID de plano de conta inválido', 'danger')
        return redirect(url_for('cadastros.plano_contas.listar'))

    connection = get_db_connection()
    if connection:
        try:
            cursor = connection.cursor(dictionary=True)

            # Buscar o status atual
            cursor.execute(
                "SELECT ativo FROM plano_contas WHERE id = %s", (id_plano,))
            plano = cursor.fetchone()

            if not plano:
                flash('Plano de conta não encontrado', 'warning')
                return redirect(url_for('cadastros.plano_contas.listar'))

            # Alternar o status com segurança para o tipo de dado
            atual = plano['ativo']
            if isinstance(atual, str) and atual.isdigit():
                atual = int(atual)
            elif not isinstance(atual, int):
                atual = 0  # valor padrão se não for um valor válido

            novo_status = 0 if atual == 1 else 1

            cursor.execute("""
                UPDATE plano_contas 
                SET ativo = %s, atualizado_em = NOW()
                WHERE id = %s
            """, (novo_status, id_plano))

            connection.commit()

            status_texto = "ativado" if novo_status == 1 else "desativado"
            flash(f'Plano de conta {status_texto} com sucesso!', 'success')

        except Exception as e:
            connection.rollback()
            logger.error(
                f"Erro ao alternar status do plano de conta: {str(e)}")
            flash(
                f'Erro ao alternar status do plano de conta: {str(e)}', 'danger')
        finally:
            cursor.close()
            connection.close()

    return redirect(url_for('cadastros.plano_contas.listar'))

# Rota para importar planos de conta a partir de um arquivo Excel


@mod_plano_contas.route('/importar', methods=['GET', 'POST'])
def importar():
    """Permite importar planos de conta a partir de um arquivo Excel"""
    # Verificar autenticação
    if 'usuario_id' not in session:
        flash('Você precisa fazer login para acessar esta página', 'warning')
        return redirect(url_for('login'))

    # Verificar permissão
    if not verificar_permissao('cadastros.plano_contas.editar'):
        flash('Você não tem permissão para importar planos de conta', 'danger')
        return redirect(url_for('cadastros.plano_contas.listar'))

    importados = []
    erros = []

    if request.method == 'POST':
        # Verificar se o arquivo foi enviado
        if 'arquivo_excel' not in request.files:
            flash('Nenhum arquivo enviado', 'warning')
            return redirect(request.url)

        arquivo = request.files['arquivo_excel']

        # Verificar se o arquivo tem nome
        if arquivo.filename == '':
            flash('Nenhum arquivo selecionado', 'warning')
            return redirect(request.url)

        # Verificar se é um arquivo válido
        if arquivo and allowed_file(arquivo.filename):
            # Salvar o arquivo temporariamente
            filename = secure_filename(arquivo.filename)
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            arquivo.save(filepath)

            substituir = 'substituir' in request.form

            try:
                # Processar o arquivo Excel com pandas
                df = pd.read_excel(filepath)

                # Verificar se as colunas necessárias existem
                colunas_necessarias = [
                    'Número', 'Índice', 'Descrição', 'Sequência']
                colunas_faltantes = [
                    col for col in colunas_necessarias if col not in df.columns]

                if colunas_faltantes:
                    flash(
                        f'Colunas obrigatórias faltantes no arquivo: {", ".join(colunas_faltantes)}', 'danger')
                    return render_template('cadastros/plano_contas/importacao/form.html',
                                           mensagem=f'Colunas obrigatórias faltantes no arquivo: {", ".join(colunas_faltantes)}',
                                           tipo_mensagem='danger')

                # Conectar ao banco de dados
                connection = get_db_connection()
                if not connection:
                    flash('Erro ao conectar ao banco de dados', 'danger')
                    return render_template('cadastros/plano_contas/importacao/form.html',
                                           mensagem='Erro ao conectar ao banco de dados',
                                           tipo_mensagem='danger')

                cursor = connection.cursor(dictionary=True)

                # Processar cada linha do DataFrame
                for index, row in df.iterrows():
                    linha = index + 2  # +2 porque o índice começa em 0 e queremos contar a linha de cabeçalho
                    try:
                        # Valores das colunas obrigatórias
                        numero = str(row['Número']).strip()
                        indice = str(row['Índice']).strip()
                        descricao = str(row['Descrição']).strip()

                        # Tentar converter sequência para int
                        try:
                            sequencia = int(row['Sequência'])
                        except (ValueError, TypeError):
                            erros.append(
                                {'linha': linha, 'mensagem': f'Sequência inválida: {row["Sequência"]}'})
                            continue

                        # Validações básicas
                        if not numero:
                            erros.append(
                                {'linha': linha, 'mensagem': 'Número é obrigatório'})
                            continue

                        if not indice:
                            erros.append(
                                {'linha': linha, 'mensagem': 'Índice é obrigatório'})
                            continue

                        if not descricao:
                            erros.append(
                                {'linha': linha, 'mensagem': 'Descrição é obrigatória'})
                            continue

                        # Verificar se já existe um plano de conta com este número ou índice
                        cursor.execute("""
                            SELECT id, numero, indice FROM plano_contas 
                            WHERE numero = %s OR indice = %s
                        """, (numero, indice))
                        plano_existente = cursor.fetchone()

                        if plano_existente:
                            if substituir:
                                # Atualizar plano de conta existente
                                cursor.execute("""
                                    UPDATE plano_contas 
                                    SET numero = %s, indice = %s, descricao = %s, sequencia = %s, atualizado_em = NOW()
                                    WHERE id = %s
                                """, (numero, indice, descricao, sequencia, plano_existente['id']))

                                importados.append({
                                    'numero': numero,
                                    'indice': indice,
                                    'descricao': descricao,
                                    'status': 'atualizado'
                                })
                            else:
                                erros.append({
                                    'linha': linha,
                                    'mensagem': f'Plano de conta já existe com número "{plano_existente["numero"]}" ou índice "{plano_existente["indice"]}". Marque a opção "Substituir" para atualizar.'
                                })
                                continue
                        else:
                            # Inserir novo plano de conta
                            cursor.execute("""
                                INSERT INTO plano_contas (numero, indice, descricao, sequencia, ativo, criado_em)
                                VALUES (%s, %s, %s, %s, TRUE, NOW())
                            """, (numero, indice, descricao, sequencia))

                            importados.append({
                                'numero': numero,
                                'indice': indice,
                                'descricao': descricao,
                                'status': 'novo'
                            })

                    except Exception as e:
                        logger.error(
                            f"Erro ao processar linha {linha}: {str(e)}")
                        erros.append({'linha': linha, 'mensagem': str(e)})

                connection.commit()
                cursor.close()
                connection.close()

                # Mensagem de sucesso
                if importados and not erros:
                    mensagem = f'Importação concluída com sucesso! {len(importados)} planos de conta processados.'
                    tipo_mensagem = 'success'
                elif importados and erros:
                    mensagem = f'Importação parcial: {len(importados)} planos de conta processados, {len(erros)} erros.'
                    tipo_mensagem = 'warning'
                else:
                    mensagem = f'Nenhum plano de conta importado. {len(erros)} erros encontrados.'
                    tipo_mensagem = 'danger'

                flash(mensagem, tipo_mensagem)

            except Exception as e:
                logger.error(f"Erro ao processar arquivo Excel: {str(e)}")
                flash(f'Erro ao processar arquivo: {str(e)}', 'danger')
                return render_template('cadastros/plano_contas/importacao/form.html',
                                       mensagem=f'Erro ao processar arquivo: {str(e)}',
                                       tipo_mensagem='danger')

            finally:
                # Remover o arquivo temporário
                if os.path.exists(filepath):
                    os.remove(filepath)

            return render_template('cadastros/plano_contas/importacao/form.html',
                                   importados=importados,
                                   erros=erros,
                                   mensagem=mensagem,
                                   tipo_mensagem=tipo_mensagem)

        else:
            flash('Tipo de arquivo não permitido. Use .xlsx ou .xls', 'warning')
            return render_template('cadastros/plano_contas/importacao/form.html',
                                   mensagem='Tipo de arquivo não permitido. Use .xlsx ou .xls',
                                   tipo_mensagem='warning')

    return render_template('cadastros/plano_contas/importacao/form.html')

# Rota para baixar modelo de arquivo Excel


@mod_plano_contas.route('/baixar-modelo')
def baixar_modelo():
    """Exibe a página de modelo para importação de planos de conta"""
    # Verificar autenticação
    if 'usuario_id' not in session:
        flash('Você precisa fazer login para acessar esta página', 'warning')
        return redirect(url_for('login'))

    return render_template('cadastros/plano_contas/importacao/modelo.html')

# Rota para baixar o arquivo XLSX de modelo


@mod_plano_contas.route('/baixar-modelo-xlsx')
def baixar_modelo_xlsx():
    """Gera e envia um arquivo Excel modelo para importação de planos de conta"""
    # Verificar autenticação
    if 'usuario_id' not in session:
        flash('Você precisa fazer login para acessar esta página', 'warning')
        return redirect(url_for('login'))

    # Criar workbook e planilha
    wb = Workbook()
    ws = wb.active
    ws.title = "Modelo Importação"

    # Definir estilos
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78",
                              end_color="1F4E78", fill_type="solid")
    header_alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Cabeçalhos
    headers = ["Número (*)", "Índice (*)", "Descrição (*)", "Sequência (*)"]
    for col_num, header_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col_num)].width = 20

    # Exemplo de dados
    example_data = [
        "1.1.01",
        "A.1.01",
        "Receitas Operacionais",
        1
    ]

    for col_num, example in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col_num, value=example)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left")

    # Salvar o workbook em um buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Enviar o arquivo para download
    return send_file(
        output,
        as_attachment=True,
        download_name="modelo_importacao_plano_contas.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# Função para inicialização do submódulo


def init_app(parent_blueprint):
    """Inicializa o submódulo plano_contas, registrando-o no blueprint pai"""
    parent_blueprint.register_blueprint(
        mod_plano_contas, url_prefix='/plano-contas')
    return parent_blueprint
