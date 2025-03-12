# coding: utf-8
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, send_file
import os
import logging
from datetime import datetime
import mysql.connector
from mysql.connector import Error
from utils.auth import verificar_permissao
import pandas as pd
from werkzeug.utils import secure_filename
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Configuração de logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('materiais.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger('materiais')

# Configuração para upload de arquivos
UPLOAD_FOLDER = os.path.join(os.path.dirname(
    os.path.abspath(__file__)), 'uploads')
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

# Criar diretório de uploads se não existir
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Verificar extensão de arquivo permitida


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# Criar o Blueprint
mod_materiais = Blueprint('materiais', __name__,
                          template_folder='templates')


def get_db_connection():
    try:
        connection = mysql.connector.connect(
            host=os.environ.get('DB_HOST', 'localhost'),
            database=os.environ.get('DB_NAME', 'sistema_solicitacoes'),
            user=os.environ.get('DB_USER', 'root'),
            password=os.environ.get('DB_PASSWORD', 'sua_senha')
        )
        return connection
    except Error as e:
        logger.error(f"Erro ao conectar ao MySQL: {e}")
        return None


# Rota principal - redirecionamento para lista de materiais
@mod_materiais.route('/')
def index():
    """Redireciona para a página de listagem de materiais"""
    return redirect(url_for('cadastros.materiais.listar'))


# Rota para listar materiais
@mod_materiais.route('/listar')
def listar():
    """Exibe a lista de materiais cadastrados"""
    # Verificar autenticação
    if 'usuario_id' not in session:
        flash('Você precisa fazer login para acessar esta página', 'warning')
        return redirect(url_for('login'))

    # Verificar permissão
    pode_editar = verificar_permissao('cadastros.materiais.editar')

    # Parâmetros de filtro
    nome = request.args.get('nome', '')
    codigo = request.args.get('codigo', '')
    status = request.args.get('status', '')
    unidade = request.args.get('unidade', '')

    materiais = []
    connection = get_db_connection()

    if connection:
        try:
            cursor = connection.cursor(dictionary=True)

            # Montagem dinâmica da consulta SQL com base nos filtros
            query = "SELECT * FROM materiais WHERE 1=1"
            params = []

            if nome:
                query += " AND nome LIKE %s"
                params.append(f"%{nome}%")

            if codigo:
                query += " AND codigo LIKE %s"
                params.append(f"%{codigo}%")

            if status and status.isdigit():
                query += " AND ativo = %s"
                params.append(int(status))

            if unidade:
                query += " AND unidade LIKE %s"
                params.append(f"%{unidade}%")

            query += " ORDER BY codigo"

            cursor.execute(query, params)
            materiais = cursor.fetchall()
            cursor.close()
        except Exception as e:
            logger.error(f"Erro ao buscar materiais: {str(e)}")
            flash(f'Erro ao buscar materiais: {str(e)}', 'danger')
        finally:
            connection.close()

    return render_template('cadastros/materiais/listar.html',
                           materiais=materiais,
                           pode_editar=pode_editar,
                           request=request)


# Rota para cadastrar novo material
@mod_materiais.route('/novo', methods=['GET', 'POST'])
def novo():
    """Exibe o formulário para cadastro de novo material e processa o envio"""
    # Verificar autenticação
    if 'usuario_id' not in session:
        flash('Você precisa fazer login para acessar esta página', 'warning')
        return redirect(url_for('login'))

    # Verificar permissão
    if not verificar_permissao('cadastros.materiais.editar'):
        flash('Você não tem permissão para cadastrar materiais', 'danger')
        return redirect(url_for('cadastros.materiais.listar'))

    form = {'errors': {}}

    if request.method == 'POST':
        codigo = request.form.get('codigo')
        nome = request.form.get('nome')
        descricao = request.form.get('descricao')
        unidade = request.form.get('unidade')
        pc = request.form.get('pc', '')
        codigo_erp = request.form.get('codigo_erp', '')

        # Validações básicas
        if not codigo:
            form['errors']['codigo'] = 'O código é obrigatório'
        if not nome:
            form['errors']['nome'] = 'O nome é obrigatório'
        if not unidade:
            form['errors']['unidade'] = 'A unidade é obrigatória'

        # Se não houver erros, prossegue com o cadastro
        if not form['errors']:
            connection = get_db_connection()
            if connection:
                try:
                    cursor = connection.cursor()
                    cursor.execute("""
                        INSERT INTO materiais (codigo, nome, descricao, unidade, pc, codigo_erp, ativo, criado_em)
                        VALUES (%s, %s, %s, %s, %s, %s, TRUE, NOW())
                    """, (codigo, nome, descricao, unidade, pc, codigo_erp))

                    connection.commit()
                    flash('Material cadastrado com sucesso!', 'success')
                    return redirect(url_for('cadastros.materiais.listar'))

                except Exception as e:
                    connection.rollback()
                    logger.error(f"Erro ao cadastrar material: {str(e)}")
                    flash(f'Erro ao cadastrar material: {str(e)}', 'danger')
                finally:
                    cursor.close()
                    connection.close()

    return render_template('cadastros/materiais/form.html', form=form)


# Rota para visualizar material
@mod_materiais.route('/visualizar/<int:id>')
def visualizar(id):
    """Exibe os detalhes de um material específico"""
    # Verificar autenticação
    if 'usuario_id' not in session:
        flash('Você precisa fazer login para acessar esta página', 'warning')
        return redirect(url_for('login'))

    # Verificar permissão de edição para ações
    pode_editar = verificar_permissao('cadastros.materiais.editar')

    # Verificar se o ID é válido
    try:
        id_material = int(id)
    except (ValueError, TypeError):
        flash('ID de material inválido', 'danger')
        return redirect(url_for('cadastros.materiais.listar'))

    material = None
    connection = get_db_connection()

    if connection:
        try:
            cursor = connection.cursor(dictionary=True)
            cursor.execute(
                "SELECT * FROM materiais WHERE id = %s", (id_material,))
            material = cursor.fetchone()
            cursor.close()

            if not material:
                flash('Material não encontrado', 'warning')
                return redirect(url_for('cadastros.materiais.listar'))

        except Exception as e:
            logger.error(f"Erro ao buscar material: {str(e)}")
            flash(f'Erro ao buscar material: {str(e)}', 'danger')
        finally:
            connection.close()

    return render_template('cadastros/materiais/visualizar.html',
                           material=material,
                           pode_editar=pode_editar)


# Rota para editar material
@mod_materiais.route('/editar/<int:id>', methods=['GET', 'POST'])
def editar(id):
    """Exibe o formulário para edição de material e processa o envio"""
    # Verificar autenticação
    if 'usuario_id' not in session:
        flash('Você precisa fazer login para acessar esta página', 'warning')
        return redirect(url_for('login'))

    # Verificar permissão
    if not verificar_permissao('cadastros.materiais.editar'):
        flash('Você não tem permissão para editar materiais', 'danger')
        return redirect(url_for('cadastros.materiais.listar'))

    # Verificar se o ID é válido
    try:
        id_material = int(id)
    except (ValueError, TypeError):
        flash('ID de material inválido', 'danger')
        return redirect(url_for('cadastros.materiais.listar'))

    material = None
    form = {'errors': {}}
    connection = get_db_connection()

    if connection:
        try:
            cursor = connection.cursor(dictionary=True)

            if request.method == 'POST':
                codigo = request.form.get('codigo')
                nome = request.form.get('nome')
                descricao = request.form.get('descricao')
                unidade = request.form.get('unidade')
                status = request.form.get('status', '1')
                pc = request.form.get('pc', '')
                codigo_erp = request.form.get('codigo_erp', '')

                # Validações básicas
                if not codigo:
                    form['errors']['codigo'] = 'O código é obrigatório'
                if not nome:
                    form['errors']['nome'] = 'O nome é obrigatório'
                if not unidade:
                    form['errors']['unidade'] = 'A unidade é obrigatória'

                # Se não houver erros, prossegue com a atualização
                if not form['errors']:
                    # Garantir que status seja um valor numérico válido
                    status_valor = 1  # valor padrão
                    if status and status.isdigit():
                        status_valor = int(status)

                    cursor.execute("""
                        UPDATE materiais 
                        SET codigo = %s, nome = %s, descricao = %s, unidade = %s, ativo = %s,
                            pc = %s, codigo_erp = %s, atualizado_em = NOW()
                        WHERE id = %s
                    """, (codigo, nome, descricao, unidade, status_valor, pc, codigo_erp, id_material))

                    connection.commit()
                    flash('Material atualizado com sucesso!', 'success')
                    return redirect(url_for('cadastros.materiais.visualizar', id=id_material))
            else:
                # Buscar dados do material para exibir no formulário
                cursor.execute(
                    "SELECT * FROM materiais WHERE id = %s", (id_material,))
                material = cursor.fetchone()

                if not material:
                    flash('Material não encontrado', 'warning')
                    return redirect(url_for('cadastros.materiais.listar'))

        except Exception as e:
            if request.method == 'POST':
                connection.rollback()
            logger.error(f"Erro ao processar material: {str(e)}")
            flash(f'Erro ao processar material: {str(e)}', 'danger')
        finally:
            cursor.close()
            connection.close()

    return render_template('cadastros/materiais/form.html', material=material, form=form)


# Rota para alternar status do material (ativar/desativar)
@mod_materiais.route('/alternar-status/<int:id>')
def alternar_status(id):
    """Alterna o status (ativo/inativo) de um material"""
    # Verificar autenticação
    if 'usuario_id' not in session:
        flash('Você precisa fazer login para acessar esta página', 'warning')
        return redirect(url_for('login'))

    # Verificar permissão
    if not verificar_permissao('cadastros.materiais.editar'):
        flash('Você não tem permissão para alterar o status de materiais', 'danger')
        return redirect(url_for('cadastros.materiais.listar'))

    # Verificar se o ID é válido
    try:
        id_material = int(id)
    except (ValueError, TypeError):
        flash('ID de material inválido', 'danger')
        return redirect(url_for('cadastros.materiais.listar'))

    connection = get_db_connection()
    if connection:
        try:
            cursor = connection.cursor(dictionary=True)

            # Buscar o status atual
            cursor.execute(
                "SELECT ativo FROM materiais WHERE id = %s", (id_material,))
            material = cursor.fetchone()

            if not material:
                flash('Material não encontrado', 'warning')
                return redirect(url_for('cadastros.materiais.listar'))

            # Alternar o status com segurança para o tipo de dado
            atual = material['ativo']
            if isinstance(atual, str) and atual.isdigit():
                atual = int(atual)
            elif not isinstance(atual, int):
                atual = 0  # valor padrão se não for um valor válido

            novo_status = 0 if atual == 1 else 1

            cursor.execute("""
                UPDATE materiais 
                SET ativo = %s, atualizado_em = NOW()
                WHERE id = %s
            """, (novo_status, id_material))

            connection.commit()

            status_texto = "ativado" if novo_status == 1 else "desativado"
            flash(f'Material {status_texto} com sucesso!', 'success')

        except Exception as e:
            connection.rollback()
            logger.error(f"Erro ao alternar status do material: {str(e)}")
            flash(f'Erro ao alternar status do material: {str(e)}', 'danger')
        finally:
            cursor.close()
            connection.close()

    return redirect(url_for('cadastros.materiais.listar'))


# Rota para importar materiais a partir de um arquivo Excel
@mod_materiais.route('/importar', methods=['GET', 'POST'])
def importar():
    """Permite importar materiais a partir de um arquivo Excel"""
    # Verificar autenticação
    if 'usuario_id' not in session:
        flash('Você precisa fazer login para acessar esta página', 'warning')
        return redirect(url_for('login'))

    # Verificar permissão
    if not verificar_permissao('cadastros.materiais.editar'):
        flash('Você não tem permissão para importar materiais', 'danger')
        return redirect(url_for('cadastros.materiais.listar'))

    importados = []
    erros = []

    if request.method == 'POST':
        # Verificar se o arquivo foi enviado
        if 'arquivo_excel' not in request.files:
            flash('Nenhum arquivo enviado', 'warning')
            return redirect(request.url)

        arquivo = request.files['arquivo_excel']

        # Verificar se o arquivo tem nome
        if arquivo.filename == '':
            flash('Nenhum arquivo selecionado', 'warning')
            return redirect(request.url)

        # Verificar se é um arquivo válido
        if arquivo and allowed_file(arquivo.filename):
            # Salvar o arquivo temporariamente
            filename = secure_filename(arquivo.filename)
            filepath = os.path.join(UPLOAD_FOLDER, filename)
            arquivo.save(filepath)

            substituir = 'substituir' in request.form

            try:
                # Processar o arquivo Excel com pandas
                df = pd.read_excel(filepath)

                # Verificar se as colunas necessárias existem
                colunas_necessarias = [
                    'Nome', 'Tipo', 'Unidade', 'Valor Unitário']
                colunas_faltantes = [
                    col for col in colunas_necessarias if col not in df.columns]

                if colunas_faltantes:
                    flash(
                        f'Colunas obrigatórias faltantes no arquivo: {", ".join(colunas_faltantes)}', 'danger')
                    return render_template('cadastros/materiais/importacao/form.html',
                                           mensagem=f'Colunas obrigatórias faltantes no arquivo: {", ".join(colunas_faltantes)}',
                                           tipo_mensagem='danger')

                # Conectar ao banco de dados
                connection = get_db_connection()
                if not connection:
                    flash('Erro ao conectar ao banco de dados', 'danger')
                    return render_template('cadastros/materiais/importacao/form.html',
                                           mensagem='Erro ao conectar ao banco de dados',
                                           tipo_mensagem='danger')

                cursor = connection.cursor(dictionary=True)

                # Processar cada linha do DataFrame
                for index, row in df.iterrows():
                    linha = index + 2  # +2 porque o índice começa em 0 e queremos contar a linha de cabeçalho
                    try:
                        # Valores das colunas obrigatórias
                        nome = str(row['Nome']).strip()
                        tipo = str(row['Tipo']).strip()
                        unidade = str(row['Unidade']).strip()

                        # Tentar converter valor para float
                        try:
                            valor_unitario = float(row['Valor Unitário'])
                        except (ValueError, TypeError):
                            erros.append(
                                {'linha': linha, 'mensagem': f'Valor unitário inválido: {row["Valor Unitário"]}'})
                            continue

                        # Gerar código automaticamente se não existir
                        codigo = f"{nome[:3].upper()}{datetime.now().strftime('%Y%m%d%H%M%S')}"

                        # Colunas opcionais
                        pc = str(row.get('PC', '')).strip() if pd.notna(
                            row.get('PC', '')) else ''
                        codigo_erp = str(row.get('Código ERP', '')).strip(
                        ) if pd.notna(row.get('Código ERP', '')) else ''

                        # Validações básicas
                        if not nome:
                            erros.append(
                                {'linha': linha, 'mensagem': 'Nome é obrigatório'})
                            continue

                        if not tipo:
                            erros.append(
                                {'linha': linha, 'mensagem': 'Tipo é obrigatório'})
                            continue

                        if not unidade:
                            erros.append(
                                {'linha': linha, 'mensagem': 'Unidade é obrigatória'})
                            continue

                        # Verificar se já existe um material com este nome ou códigos
                        cursor.execute("""
                            SELECT id, nome FROM materiais 
                            WHERE nome = %s OR 
                                  (pc = %s AND pc != '') OR 
                                  (codigo_erp = %s AND codigo_erp != '')
                        """, (nome, pc, codigo_erp))
                        material_existente = cursor.fetchone()

                        if material_existente:
                            if substituir:
                                # Atualizar material existente
                                cursor.execute("""
                                    UPDATE materiais 
                                    SET tipo = %s, unidade = %s, valor_unitario = %s, 
                                        pc = %s, codigo_erp = %s, atualizado_em = NOW()
                                    WHERE id = %s
                                """, (tipo, unidade, valor_unitario, pc, codigo_erp, material_existente['id']))

                                importados.append({
                                    'nome': nome,
                                    'pc': pc,
                                    'codigo_erp': codigo_erp,
                                    'status': 'atualizado'
                                })
                            else:
                                erros.append({
                                    'linha': linha,
                                    'mensagem': f'Material já existe com nome "{material_existente["nome"]}". Marque a opção "Substituir" para atualizar.'
                                })
                                continue
                        else:
                            # Inserir novo material
                            cursor.execute("""
                                INSERT INTO materiais (codigo, nome, tipo, unidade, valor_unitario, pc, codigo_erp, ativo, criado_em)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, TRUE, NOW())
                            """, (codigo, nome, tipo, unidade, valor_unitario, pc, codigo_erp))

                            importados.append({
                                'nome': nome,
                                'pc': pc,
                                'codigo_erp': codigo_erp,
                                'status': 'novo'
                            })

                    except Exception as e:
                        logger.error(
                            f"Erro ao processar linha {linha}: {str(e)}")
                        erros.append({'linha': linha, 'mensagem': str(e)})

                connection.commit()
                cursor.close()
                connection.close()

                # Mensagem de sucesso
                if importados and not erros:
                    mensagem = f'Importação concluída com sucesso! {len(importados)} materiais processados.'
                    tipo_mensagem = 'success'
                elif importados and erros:
                    mensagem = f'Importação parcial: {len(importados)} materiais processados, {len(erros)} erros.'
                    tipo_mensagem = 'warning'
                else:
                    mensagem = f'Nenhum material importado. {len(erros)} erros encontrados.'
                    tipo_mensagem = 'danger'

                flash(mensagem, tipo_mensagem)

            except Exception as e:
                logger.error(f"Erro ao processar arquivo Excel: {str(e)}")
                flash(f'Erro ao processar arquivo: {str(e)}', 'danger')
                return render_template('cadastros/materiais/importacao/form.html',
                                       mensagem=f'Erro ao processar arquivo: {str(e)}',
                                       tipo_mensagem='danger')

            finally:
                # Remover o arquivo temporário
                if os.path.exists(filepath):
                    os.remove(filepath)

            return render_template('cadastros/materiais/importacao/form.html',
                                   importados=importados,
                                   erros=erros,
                                   mensagem=mensagem,
                                   tipo_mensagem=tipo_mensagem)

        else:
            flash('Tipo de arquivo não permitido. Use .xlsx ou .xls', 'warning')
            return render_template('cadastros/materiais/importacao/form.html',
                                   mensagem='Tipo de arquivo não permitido. Use .xlsx ou .xls',
                                   tipo_mensagem='warning')

    return render_template('cadastros/materiais/importacao/form.html')


# Rota para baixar modelo de arquivo Excel
@mod_materiais.route('/baixar-modelo')
def baixar_modelo():
    """Exibe a página de modelo para importação de materiais"""
    # Verificar autenticação
    if 'usuario_id' not in session:
        flash('Você precisa fazer login para acessar esta página', 'warning')
        return redirect(url_for('login'))

    return render_template('cadastros/materiais/importacao/modelo.html')


# Rota para baixar o arquivo XLSX de modelo
@mod_materiais.route('/baixar-modelo-xlsx')
def baixar_modelo_xlsx():
    """Gera e envia um arquivo Excel modelo para importação de materiais"""
    # Verificar autenticação
    if 'usuario_id' not in session:
        flash('Você precisa fazer login para acessar esta página', 'warning')
        return redirect(url_for('login'))

    # Criar workbook e planilha
    wb = Workbook()
    ws = wb.active
    ws.title = "Modelo Importação"

    # Definir estilos
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78",
                              end_color="1F4E78", fill_type="solid")
    header_alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Cabeçalhos
    headers = ["Nome (*)", "Tipo (*)", "Unidade (*)",
               "Valor Unitário (*)", "PC", "Código ERP"]
    for col_num, header_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col_num)].width = 20

    # Exemplo de dados
    example_data = [
        "Parafuso Allen M10",
        "Fixação",
        "UN",
        2.50,
        "PC001",
        "ERP001"
    ]

    for col_num, example in enumerate(example_data, 1):
        cell = ws.cell(row=2, column=col_num, value=example)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left")

    # Salvar o workbook em um buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Enviar o arquivo para download
    return send_file(
        output,
        as_attachment=True,
        download_name="modelo_importacao_materiais.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# Função para inicialização do submódulo
def init_app(parent_blueprint):
    """Inicializa o submódulo materiais, registrando-o no blueprint pai"""
    # parent_blueprint.register_blueprint(mod_materiais, url_prefix='/materiais')
    return parent_blueprint
